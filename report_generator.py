"""
Report generation functions for the Amazon Sales Automation project.
Generates a formatted Excel report and a PDF summary report with charts.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

REPORTS_DIR = "reports"

HEADER_FILL = PatternFill(start_color="4C72B0", end_color="4C72B0", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F1F1F")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize(ws, max_width=45):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_width, max(10, length + 2))


# ---------------------------------------------------------------------------
# EXCEL REPORT
# ---------------------------------------------------------------------------
def generate_excel_report(clean_df, stats, dq, out_dir=REPORTS_DIR):
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = "Amazon Sales Dataset - Data Cleaning & Reporting Automation"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:D2")

    summary_rows = [
        ("Metric", "Value"),
        ("Total Products (cleaned)", stats["total_products"]),
        ("Total Categories", stats["total_categories"]),
        ("Average Rating", stats["average_rating"]),
        ("Highest Rating", stats["highest_rating"]),
        ("Lowest Rating", stats["lowest_rating"]),
        ("Average Discount %", stats["average_discount_pct"]),
        ("Highest Discount %", stats["highest_discount_pct"]),
        ("Average Discounted Price (₹)", stats["average_discounted_price"]),
        ("Average Actual Price (₹)", stats["average_actual_price"]),
        ("Raw Rows (before cleaning)", dq["n_rows"]),
        ("Duplicate Rows Removed", dq["duplicate_rows"]),
    ]
    start_row = 4
    for i, (label, value) in enumerate(summary_rows):
        r = start_row + i
        ws.cell(row=r, column=2, value=label).font = BODY_FONT if i else HEADER_FONT
        ws.cell(row=r, column=3, value=value).font = BODY_FONT
        if i == 0:
            _style_header_row(ws, r, 3)
            for c in (2, 3):
                ws.cell(row=r, column=c).fill = HEADER_FILL
                ws.cell(row=r, column=c).font = HEADER_FONT
    _autosize(ws)

    # --- Sheet 2: Cleaned Data ---
    ws2 = wb.create_sheet("Cleaned Data")
    export_cols = [c for c in clean_df.columns if c not in ("about_product", "img_link")]
    export_df = clean_df[export_cols]
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), start=1):
        ws2.append(row)
    _style_header_row(ws2, 1, len(export_cols))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(export_cols)):
        for cell in row:
            cell.font = BODY_FONT
    ws2.freeze_panes = "A2"
    _autosize(ws2, max_width=30)

    # --- Sheet 3: Category Statistics ---
    ws3 = wb.create_sheet("Category Stats")
    ws3.append(["Category", "Number of Products"])
    _style_header_row(ws3, 1, 2)
    for cat, count in stats["top_categories"].items():
        ws3.append([cat, int(count)])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.font = BODY_FONT
    _autosize(ws3)

    # Add a bar chart for category stats
    chart = BarChart()
    chart.title = "Top Categories by Number of Products"
    chart.y_axis.title = "Number of Products"
    chart.x_axis.title = "Category"
    data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    ws3.add_chart(chart, "D2")

    # --- Sheet 4: Top Rated / Most Reviewed / Most Expensive ---
    ws4 = wb.create_sheet("Top Products")
    row_ptr = 1
    for title, table in [
        ("Top Rated Products", stats["top_rated_products"]),
        ("Most Reviewed Products", stats["most_reviewed_products"]),
        ("Most Expensive Products", stats["most_expensive_products"]),
    ]:
        ws4.cell(row=row_ptr, column=1, value=title).font = Font(name="Arial", bold=True, size=12)
        row_ptr += 1
        header_row = row_ptr
        for j, col in enumerate(table.columns, start=1):
            ws4.cell(row=row_ptr, column=j, value=col)
        _style_header_row(ws4, header_row, len(table.columns))
        row_ptr += 1
        for _, r in table.iterrows():
            for j, val in enumerate(r, start=1):
                ws4.cell(row=row_ptr, column=j, value=val).font = BODY_FONT
            row_ptr += 1
        row_ptr += 2  # blank rows between tables
    _autosize(ws4, max_width=45)

    out_path = os.path.join(out_dir, "report.xlsx")
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# PDF REPORT
# ---------------------------------------------------------------------------
def generate_pdf_report(stats, dq, chart_paths, out_dir=REPORTS_DIR):
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "report.pdf")

    doc = SimpleDocTemplate(out_path, pagesize=letter,
                             topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontSize=20)
    heading_style = styles["Heading2"]
    body_style = styles["Normal"]

    story = []

    # --- Title Page ---
    story.append(Paragraph("Amazon Sales Dataset", title_style))
    story.append(Paragraph("Data Cleaning & Reporting Automation Project", styles["Heading3"]))
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "This report documents an end-to-end data cleaning, exploratory analysis, and "
        "reporting automation pipeline built on the Amazon Sales product dataset "
        "(1,465 raw rows, 16 columns), prepared as part of a Data Analysis internship task.",
        body_style,
    ))
    story.append(Spacer(1, 20))

    # --- Dataset Summary ---
    story.append(Paragraph("Dataset Summary", heading_style))
    summary_data = [
        ["Metric", "Value"],
        ["Raw rows", dq["n_rows"]],
        ["Duplicate rows removed", dq["duplicate_rows"]],
        ["Cleaned rows", stats["total_products"]],
        ["Total categories", stats["total_categories"]],
        ["Average rating", stats["average_rating"]],
        ["Average discount %", f"{stats['average_discount_pct']}%"],
        ["Average discounted price", f"₹{stats['average_discounted_price']}"],
        ["Average actual price", f"₹{stats['average_actual_price']}"],
    ]
    t = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C72B0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    # --- Key Findings ---
    story.append(Paragraph("Key Findings", heading_style))
    top_cat_name = stats["top_categories"].index[0]
    most_reviewed_name = stats["most_reviewed_products"].iloc[0]["product_name"]
    most_expensive_name = stats["most_expensive_products"].iloc[0]["product_name"]
    findings = [
        f"The largest product category is <b>{top_cat_name}</b>, "
        f"with {int(stats['top_categories'].iloc[0])} listed products.",
        f"Average customer rating across all products is <b>{stats['average_rating']}</b> "
        f"out of 5, with ratings ranging from {stats['lowest_rating']} to {stats['highest_rating']}.",
        f"Products carry an average discount of <b>{stats['average_discount_pct']}%</b>, "
        f"with the steepest discount reaching {stats['highest_discount_pct']}%.",
        f"The most-reviewed product is <b>{most_reviewed_name[:60]}</b>.",
        f"The most expensive product listed is <b>{most_expensive_name[:60]}</b> "
        f"at ₹{stats['most_expensive_products'].iloc[0]['actual_price']:,.0f}.",
    ]
    for f in findings:
        story.append(Paragraph("• " + f, body_style))
        story.append(Spacer(1, 4))

    story.append(PageBreak())

    # --- Charts ---
    story.append(Paragraph("Charts & Visualizations", heading_style))
    story.append(Spacer(1, 8))
    chart_titles = {
        "category_chart.png": "Top 10 Categories by Number of Listings",
        "rating_chart.png": "Rating Distribution",
        "top_products.png": "Top 10 Most Reviewed Products",
        "discount_chart.png": "Discount Percentage Distribution",
        "avg_price_by_category.png": "Average Price by Category",
        "top_rated_categories.png": "Top Rated Categories",
        "price_scatter.png": "Actual Price vs Discounted Price",
        "expensive_products.png": "Top 10 Most Expensive Products",
    }
    for i, path in enumerate(chart_paths):
        fname = os.path.basename(path)
        story.append(Paragraph(chart_titles.get(fname, fname), styles["Heading4"]))
        story.append(Image(path, width=6.2 * inch, height=3.7 * inch))
        story.append(Spacer(1, 10))
        if i % 2 == 1 and i != len(chart_paths) - 1:
            story.append(PageBreak())

    story.append(PageBreak())

    # --- Conclusion ---
    story.append(Paragraph("Conclusion", heading_style))
    story.append(Paragraph(
        "The raw Amazon sales dataset contained duplicate rows, currency symbols and "
        "thousands separators embedded in numeric fields, inconsistent text spacing, "
        "and a small number of invalid rating values. After cleaning, the dataset was "
        "reduced to a consistent, analysis-ready table. The automated pipeline (main.py) "
        "can be re-run on any updated export of this dataset to regenerate the cleaned "
        "CSV, charts, and this report without manual intervention.",
        body_style,
    ))

    doc.build(story)
    return out_path
